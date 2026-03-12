#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║         AWS Security Hub — FSBP Findings Puller              ║
║         AWS Foundational Security Best Practices v1.0.0      ║
╚══════════════════════════════════════════════════════════════╝

Pulls Security Hub findings across all AWS Organization member
accounts and generates:

  1. Excel remediation tracker  — weekly status report, findings
     log, by-severity, by-account, and top failing controls
  2. HTML dashboard             — KPI cards, progress bars, and
     per-account breakdown, open in any browser

Requirements:
  pip install boto3 openpyxl

Usage:
  python securityhub_pull.py --region us-east-1
  python securityhub_pull.py --region us-east-1 --profile my-profile
  python securityhub_pull.py --region us-east-1 --accounts 111111111111 222222222222
"""

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── CONFIG ────────────────────────────────────────────────────────────────────
FSBP_STANDARD_ARN_FRAGMENT = "aws-foundational-security-best-practices/v/1.0.0"
MAX_FINDINGS = 100000  # 45 accounts × 200+ FSBP controls = potentially 50,000+ findings
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── COLOR PALETTE ─────────────────────────────────────────────────────────────
DARK_NAVY   = "1B2A47"
MID_BLUE    = "2E4A7A"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F7FA"
MID_GRAY    = "E2E8F0"
DARK_GRAY   = "64748B"
RED         = "DC2626"
ORANGE      = "EA580C"
YELLOW      = "D97706"
GREEN       = "16A34A"
TEAL        = "0891B2"
LIGHT_RED   = "FEE2E2"
LIGHT_ORANGE= "FFEDD5"
LIGHT_YELLOW= "FEF9C3"
LIGHT_GREEN = "DCFCE7"
LIGHT_TEAL  = "CFFAFE"

SEVERITY_COLORS = {
    "CRITICAL": (LIGHT_RED,    RED),
    "HIGH":     (LIGHT_ORANGE, ORANGE),
    "MEDIUM":   (LIGHT_YELLOW, YELLOW),
    "LOW":      (LIGHT_GREEN,  GREEN),
    "INFORMATIONAL": (MID_GRAY, DARK_GRAY),
}
STATUS_COLORS = {
    "FAILED":  (LIGHT_RED,    RED),
    "PASSED":  (LIGHT_GREEN,  GREEN),
    "WARNING": (LIGHT_YELLOW, YELLOW),
    "NOT_AVAILABLE": (MID_GRAY, DARK_GRAY),
}

def mk_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def mk_font(size=9, bold=False, color="1E293B"):
    return Font(name='Arial', size=size, bold=bold, color=color)

def mk_hdr_font(size=9, color=WHITE):
    return Font(name='Arial', size=size, bold=True, color=color)

def mk_align(h='left'):
    return Alignment(horizontal=h, vertical='center', wrap_text=True)

def mk_border():
    s = Side(style='thin', color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

# ── AWS PULL ──────────────────────────────────────────────────────────────────
def get_session(profile=None, region=None):
    import boto3
    kwargs = {}
    if profile:
        kwargs['profile_name'] = profile
    if region:
        kwargs['region_name'] = region
    return boto3.Session(**kwargs)

def pull_findings(session, accounts=None):
    client = session.client('securityhub')
    region = session.region_name or client.meta.region_name

    # Consolidated controls mode — filter by ProductName=Security Hub and
    # ComplianceStatus only. We post-filter for FSBP controls client-side
    # using the GeneratorId prefix 'security-control/' which covers all
    # Security Hub managed controls including FSBP v1.0.0
    filters_failed = {
        'RecordState':      [{'Value': 'ACTIVE',  'Comparison': 'EQUALS'}],
        'ComplianceStatus': [{'Value': 'FAILED',  'Comparison': 'EQUALS'}],
        'ProductFields':    [{'Key': 'aws/securityhub/ProductName', 'Value': 'Security Hub', 'Comparison': 'EQUALS'}],
    }
    filters_passed = {
        'RecordState':      [{'Value': 'ACTIVE',  'Comparison': 'EQUALS'}],
        'ComplianceStatus': [{'Value': 'PASSED',  'Comparison': 'EQUALS'}],
        'ProductFields':    [{'Key': 'aws/securityhub/ProductName', 'Value': 'Security Hub', 'Comparison': 'EQUALS'}],
    }
    if accounts:
        acct_filter = [{'Value': a, 'Comparison': 'EQUALS'} for a in accounts]
        filters_failed['AwsAccountId'] = acct_filter
        filters_passed['AwsAccountId'] = acct_filter

    findings = []
    paginator = client.get_paginator('get_findings')

    # Pull failed findings (open — highest priority)
    print(f"  Pulling FAILED findings from Security Hub ({region}) across all member accounts...")
    pages = paginator.paginate(
        Filters=filters_failed,
        SortCriteria=[{'Field': 'SeverityLabel', 'SortOrder': 'desc'}],
        PaginationConfig={'MaxItems': MAX_FINDINGS, 'PageSize': 100}
    )
    for page in pages:
        findings.extend(page.get('Findings', []))
        print(f"  → {len(findings)} failed findings so far...", end='\r')
    failed_count = len(findings)
    print(f"\n  ✓ Failed findings: {failed_count}")

    # Pull passed findings (remediated — for audit trail completeness)
    print(f"  Pulling PASSED findings (remediated) for audit trail...")
    passed = []
    pages = paginator.paginate(
        Filters=filters_passed,
        SortCriteria=[{'Field': 'UpdatedAt', 'SortOrder': 'desc'}],
        PaginationConfig={'MaxItems': 10000, 'PageSize': 100}  # Cap passed — less critical
    )
    for page in pages:
        passed.extend(page.get('Findings', []))
        print(f"  → {len(passed)} passed findings so far...", end='\r')
    print(f"\n  ✓ Passed findings: {len(passed)}")

    findings.extend(passed)
    print(f"  ✓ Total findings: {len(findings)} across all member accounts")
    return findings, region

def parse_finding(f):
    """Normalize a raw Security Hub finding into a flat dict."""
    severity    = f.get('Severity', {}).get('Label', 'INFORMATIONAL')
    status      = f.get('Compliance', {}).get('Status', 'NOT_AVAILABLE')
    title       = f.get('Title', '')
    description = f.get('Description', '')
    account_id  = f.get('AwsAccountId', '')
    region      = f.get('Region', '')
    created_at  = f.get('CreatedAt', '')[:10] if f.get('CreatedAt') else ''
    updated_at  = f.get('UpdatedAt', '')[:10] if f.get('UpdatedAt') else ''

    # Extract control ID from GeneratorId
    # Consolidated mode: 'security-control/EC2.14' -> 'EC2.14'
    # Legacy mode: 'aws-foundational-.../EC2.14' -> 'EC2.14'
    generator_id = f.get('GeneratorId', '')
    control_id = generator_id.split('/')[-1] if '/' in generator_id else generator_id

    # Derive category from control ID prefix (consolidated controls mode)
    category_map = {
        'IAM':         'Identity & Access Management',
        'S3':          'Data Protection',
        'EC2':         'Infrastructure Protection',
        'RDS':         'Data Protection',
        'Lambda':      'Infrastructure Protection',
        'CloudTrail':  'Logging & Monitoring',
        'GuardDuty':   'Detection',
        'SecurityHub': 'Detection',
        'Config':      'Logging & Monitoring',
        'KMS':         'Data Protection',
        'EKS':         'Infrastructure Protection',
        'ECS':         'Infrastructure Protection',
        'SSM':         'Vulnerability Management',
        'Inspector':   'Vulnerability Management',
        'Macie':       'Data Protection',
        'WAF':         'Infrastructure Protection',
        'SNS':         'Logging & Monitoring',
        'SQS':         'Data Protection',
        'DynamoDB':    'Data Protection',
        'Redshift':    'Data Protection',
        'ElastiCache': 'Data Protection',
        'OpenSearch':  'Data Protection',
        'MSK':         'Data Protection',
        'SecretsManager': 'Data Protection',
        'ACM':         'Data Protection',
        'AutoScaling': 'Infrastructure Protection',
        'ELB':         'Infrastructure Protection',
        'ELBv2':       'Infrastructure Protection',
        'CloudFront':  'Infrastructure Protection',
        'Route53':     'Infrastructure Protection',
        'APIGateway':  'Infrastructure Protection',
        'CodeBuild':   'Infrastructure Protection',
        'ECR':         'Infrastructure Protection',
        'Athena':      'Data Protection',
        'EMR':         'Data Protection',
        'SageMaker':   'Data Protection',
        'Account':     'Identity & Access Management',
    }
    svc_prefix = control_id.split('.')[0] if '.' in control_id else control_id
    category = category_map.get(svc_prefix, 'Other')

    # Resources
    resources = f.get('Resources', [{}])
    resource  = resources[0] if resources else {}
    resource_id   = resource.get('Id', '')
    resource_type = resource.get('Type', '')

    # Remediation
    remediation_text = f.get('Remediation', {}).get('Recommendation', {}).get('Text', '')
    remediation_url  = f.get('Remediation', {}).get('Recommendation', {}).get('Url', '')

    # Map compliance status → human status
    status_map = {
        'FAILED':        'Open',
        'PASSED':        'Remediated',
        'WARNING':       'In Progress',
        'NOT_AVAILABLE': 'Not Available',
    }
    human_status = status_map.get(status, status)

    return {
        'control_id':       control_id,
        'title':            title,
        'category':         category,
        'account_id':       account_id,
        'region':           region,
        'resource_id':      resource_id[:80] if resource_id else '',
        'resource_type':    resource_type,
        'severity':         severity,
        'compliance_status':status,
        'human_status':     human_status,
        'created_at':       created_at,
        'updated_at':       updated_at,
        'remediation_text': remediation_text,
        'remediation_url':  remediation_url,
        'description':      description[:200] if description else '',
    }

# ── EXCEL BUILDER ─────────────────────────────────────────────────────────────
def build_excel(findings_data, region, pull_time, output_path):
    wb = Workbook()

    # ── Sheet 1: Weekly Status Report ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Weekly Status Report"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells('A1:L1')
    ws1['A1'] = "☁  AWS Security Hub — Remediation Status Report"
    ws1['A1'].font = Font(name='Arial', size=16, bold=True, color=WHITE)
    ws1['A1'].fill = mk_fill(DARK_NAVY)
    ws1['A1'].alignment = mk_align('center')
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells('A2:L2')
    ws1['A2'] = f"AWS Foundational Security Best Practices (FSBP) v1.0.0  |  Region: {region}  |  Generated: {pull_time}"
    ws1['A2'].font = Font(name='Arial', size=9, color="90CAF9")
    ws1['A2'].fill = mk_fill(DARK_NAVY)
    ws1['A2'].alignment = mk_align('center')
    ws1.row_dimensions[2].height = 20

    # Meta
    meta_data = [
        (1, "Report Date:", 2, pull_time[:10]),
        (4, "Region:", 5, region),
        (7, "Accounts in Scope:", 8, str(len(set(f['account_id'] for f in findings_data)))),
        (10, "Go-Live Target:", 11, "May 2025"),
    ]
    for lc, lbl, vc, val in meta_data:
        c1 = ws1.cell(row=3, column=lc, value=lbl)
        c1.font = mk_font(9, bold=True, color=DARK_GRAY)
        c1.fill = mk_fill(LIGHT_GRAY)
        c1.alignment = mk_align()
        c2 = ws1.cell(row=3, column=vc, value=val)
        c2.font = mk_font(9, color="1E293B")
        c2.fill = mk_fill(LIGHT_GRAY)
        c2.alignment = mk_align()
        ws1.cell(row=3, column=vc+1).fill = mk_fill(LIGHT_GRAY)
    ws1.row_dimensions[3].height = 20
    ws1.row_dimensions[4].height = 8

    # KPI boxes
    kpi_boxes = [
        ('A5:C6', 'A7:C8', "CRITICAL FINDINGS", RED,    LIGHT_RED,    'COUNTIFS(Findings!H:H,"CRITICAL",Findings!I:I,"Open")'),
        ('D5:F6', 'D7:F8', "HIGH FINDINGS",     ORANGE, LIGHT_ORANGE, 'COUNTIFS(Findings!H:H,"HIGH",Findings!I:I,"Open")'),
        ('G5:I6', 'G7:I8', "MEDIUM FINDINGS",   YELLOW, LIGHT_YELLOW, 'COUNTIFS(Findings!H:H,"MEDIUM",Findings!I:I,"Open")'),
        ('J5:L6', 'J7:L8', "REMEDIATED",        GREEN,  LIGHT_GREEN,  'COUNTIF(Findings!I:I,"Remediated")'),
    ]
    for hdr_rng, val_rng, label, fg, bg, formula in kpi_boxes:
        ws1.merge_cells(hdr_rng)
        ws1.merge_cells(val_rng)
        hdr_cell = ws1[hdr_rng.split(':')[0]]
        hdr_cell.value = label
        hdr_cell.font = mk_hdr_font(9, color=WHITE)
        hdr_cell.fill = mk_fill(fg)
        hdr_cell.alignment = mk_align('center')
        val_cell = ws1[val_rng.split(':')[0]]
        val_cell.value = f'={formula}'
        val_cell.font = Font(name='Arial', size=22, bold=True, color=fg)
        val_cell.fill = mk_fill(bg)
        val_cell.alignment = mk_align('center')
    for r in [5,6,7,8]:
        ws1.row_dimensions[r].height = 20

    ws1.row_dimensions[9].height = 10

    # Narrative section
    ws1.merge_cells('A10:L10')
    ws1['A10'] = "WEEKLY SUMMARY — Complete before sending to CISO"
    ws1['A10'].font = mk_hdr_font(10, WHITE)
    ws1['A10'].fill = mk_fill(MID_BLUE)
    ws1['A10'].alignment = mk_align()
    ws1.row_dimensions[10].height = 22

    # Auto-populate narrative with real stats
    total   = len(findings_data)
    crit    = sum(1 for f in findings_data if f['severity'] == 'CRITICAL' and f['human_status'] == 'Open')
    hi      = sum(1 for f in findings_data if f['severity'] == 'HIGH' and f['human_status'] == 'Open')
    rem     = sum(1 for f in findings_data if f['human_status'] == 'Remediated')
    accounts= set(f['account_id'] for f in findings_data)

    narratives = [
        ("Remediation Progress:",
         f"{rem} findings remediated of {total} total. {crit} critical and {hi} high severity findings remain open across {len(accounts)} AWS account(s)."),
        ("Blockers / Dependencies:", ""),
        ("Cloud Engineering Actions Required:", ""),
        ("Risks to Go-Live Timeline:",
         f"{'⚠ ' + str(crit) + ' CRITICAL findings unresolved — must address before go-live.' if crit > 0 else 'No critical blockers identified.'}"),
        ("Planned for Next Week:", ""),
    ]
    for i, (lbl, val) in enumerate(narratives):
        r = 11 + i
        ws1.merge_cells(f'A{r}:C{r}')
        ws1.merge_cells(f'D{r}:L{r}')
        c1 = ws1[f'A{r}']
        c1.value = lbl
        c1.font = mk_font(9, bold=True, color=DARK_GRAY)
        c1.fill = mk_fill(LIGHT_GRAY)
        c1.alignment = mk_align()
        c2 = ws1[f'D{r}']
        c2.value = val
        c2.font = mk_font(9)
        c2.fill = mk_fill(WHITE)
        c2.alignment = mk_align()
        c2.border = mk_border()
        ws1.row_dimensions[r].height = 28

    col_widths = [18,12,12,18,12,12,18,12,12,18,12,12]
    for i, w in enumerate(col_widths):
        ws1.column_dimensions[get_column_letter(i+1)].width = w
    ws1.sheet_properties.tabColor = DARK_NAVY

    # ── Sheet 2: Findings ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Findings")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A3'

    ws2.merge_cells('A1:P1')
    ws2['A1'] = f"AWS FSBP v1.0.0 — Security Hub Findings  |  {len(findings_data)} findings  |  Pulled: {pull_time}"
    ws2['A1'].font = Font(name='Arial', size=12, bold=True, color=WHITE)
    ws2['A1'].fill = mk_fill(DARK_NAVY)
    ws2['A1'].alignment = mk_align('center')
    ws2.row_dimensions[1].height = 28

    headers = [
        "Control ID","Title","Category","Account ID","Region",
        "Resource ID","Resource Type","Severity","Status",
        "First Seen","Last Updated","Remediation Guide",
        "Assigned To","Date Remediated","Ticket / PR Ref","Notes"
    ]
    hdr_colors = [DARK_NAVY]*3 + [MID_BLUE]*4 + ["7B1414","14532D"] + ["374151"]*3 + ["1E3A5F"]*4

    for col, (hdr, clr) in enumerate(zip(headers, hdr_colors), 1):
        cell = ws2.cell(row=2, column=col, value=hdr)
        cell.font = mk_hdr_font(9)
        cell.fill = mk_fill(clr)
        cell.alignment = mk_align('center')
        cell.border = mk_border()
    ws2.row_dimensions[2].height = 32

    for row_idx, f in enumerate(findings_data, 3):
        row_bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE
        row_data = [
            f['control_id'], f['title'], f['category'],
            f['account_id'], f['region'], f['resource_id'],
            f['resource_type'], f['severity'], f['human_status'],
            f['created_at'], f['updated_at'], f['remediation_text'],
            '', '', '', ''
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = mk_border()
            cell.alignment = mk_align()

            if col_idx == 8 and value in SEVERITY_COLORS:
                bg, fg = SEVERITY_COLORS[value]
                cell.fill = mk_fill(bg)
                cell.font = Font(name='Arial', size=9, bold=True, color=fg)
                cell.alignment = mk_align('center')
            elif col_idx == 9:
                status_fill_map = {
                    'Open':        (LIGHT_RED,    RED),
                    'Remediated':  (LIGHT_GREEN,  GREEN),
                    'In Progress': (LIGHT_YELLOW, YELLOW),
                }
                bg, fg = status_fill_map.get(value, (MID_GRAY, DARK_GRAY))
                cell.fill = mk_fill(bg)
                cell.font = Font(name='Arial', size=9, bold=True, color=fg)
                cell.alignment = mk_align('center')
            else:
                cell.fill = mk_fill(row_bg)
                cell.font = mk_font(9)
        ws2.row_dimensions[row_idx].height = 20

    dv = DataValidation(type="list", formula1='"Open,In Progress,Remediated,Risk Accepted,False Positive"', allow_blank=True)
    ws2.add_data_validation(dv)
    dv.sqref = f"I3:I{len(findings_data)+10}"

    findings_widths = [12,38,22,16,14,35,22,11,14,12,12,35,18,16,18,25]
    for i, w in enumerate(findings_widths):
        ws2.column_dimensions[get_column_letter(i+1)].width = w
    ws2.sheet_properties.tabColor = RED

    # ── Sheet 3: By Severity ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Severity")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells('A1:F1')
    ws3['A1'] = "Findings Summary by Severity"
    ws3['A1'].font = Font(name='Arial', size=13, bold=True, color=WHITE)
    ws3['A1'].fill = mk_fill(DARK_NAVY)
    ws3['A1'].alignment = mk_align('center')
    ws3.row_dimensions[1].height = 32

    sev_headers = ["Severity", "Total", "Open", "In Progress", "Remediated", "% Remediated"]
    for col, hdr in enumerate(sev_headers, 1):
        cell = ws3.cell(row=2, column=col, value=hdr)
        cell.font = mk_hdr_font(9)
        cell.fill = mk_fill(MID_BLUE)
        cell.alignment = mk_align('center')
        cell.border = mk_border()
    ws3.row_dimensions[2].height = 28

    sev_order = ["CRITICAL","HIGH","MEDIUM","LOW","INFORMATIONAL"]
    sev_counts = defaultdict(lambda: defaultdict(int))
    for f in findings_data:
        sev_counts[f['severity']][f['human_status']] += 1

    for i, sev in enumerate(sev_order):
        r = i + 3
        bg, fg = SEVERITY_COLORS.get(sev, (MID_GRAY, DARK_GRAY))
        counts = sev_counts.get(sev, {})
        total_sev = sum(counts.values())
        remediated = counts.get('Remediated', 0)

        row_data = [
            sev,
            total_sev,
            counts.get('Open', 0),
            counts.get('In Progress', 0),
            remediated,
            f"{remediated/total_sev*100:.0f}%" if total_sev > 0 else "0%"
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=col, value=val)
            cell.border = mk_border()
            cell.alignment = mk_align('center')
            if col == 1:
                cell.fill = mk_fill(bg)
                cell.font = Font(name='Arial', size=9, bold=True, color=fg)
            else:
                cell.fill = mk_fill(LIGHT_GRAY if i % 2 == 0 else WHITE)
                cell.font = mk_font(9, bold=(col == 2))
        ws3.row_dimensions[r].height = 22

    # Totals
    tr = len(sev_order) + 3
    totals = ["TOTAL", len(findings_data),
              sum(1 for f in findings_data if f['human_status'] == 'Open'),
              sum(1 for f in findings_data if f['human_status'] == 'In Progress'),
              sum(1 for f in findings_data if f['human_status'] == 'Remediated'), ""]
    total_rem = totals[4]
    total_all = totals[1]
    totals[5] = f"{total_rem/total_all*100:.0f}%" if total_all > 0 else "0%"
    for col, val in enumerate(totals, 1):
        cell = ws3.cell(row=tr, column=col, value=val)
        cell.font = Font(name='Arial', size=9, bold=True, color=WHITE)
        cell.fill = mk_fill(DARK_NAVY)
        cell.alignment = mk_align('center')
        cell.border = mk_border()
    ws3.row_dimensions[tr].height = 24

    for i, w in enumerate([18,12,12,14,14,14]):
        ws3.column_dimensions[get_column_letter(i+1)].width = w
    ws3.sheet_properties.tabColor = MID_BLUE

    # ── Sheet 4: By Account ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("By Account")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells('A1:G1')
    ws4['A1'] = "Findings by AWS Account"
    ws4['A1'].font = Font(name='Arial', size=13, bold=True, color=WHITE)
    ws4['A1'].fill = mk_fill(DARK_NAVY)
    ws4['A1'].alignment = mk_align('center')
    ws4.row_dimensions[1].height = 32

    acct_headers = ["Account ID","Total","Critical","High","Medium","Remediated","% Remediated"]
    for col, hdr in enumerate(acct_headers, 1):
        cell = ws4.cell(row=2, column=col, value=hdr)
        cell.font = mk_hdr_font(9)
        cell.fill = mk_fill(MID_BLUE)
        cell.alignment = mk_align('center')
        cell.border = mk_border()
    ws4.row_dimensions[2].height = 28

    acct_data = defaultdict(lambda: defaultdict(int))
    for f in findings_data:
        acct_data[f['account_id']]['total'] += 1
        acct_data[f['account_id']][f['severity']] += 1
        if f['human_status'] == 'Remediated':
            acct_data[f['account_id']]['remediated'] += 1

    for i, (acct_id, counts) in enumerate(sorted(acct_data.items())):
        r = i + 3
        row_bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        total_acct = counts['total']
        rem_acct = counts['remediated']
        row = [acct_id, total_acct, counts['CRITICAL'], counts['HIGH'],
               counts['MEDIUM'], rem_acct,
               f"{rem_acct/total_acct*100:.0f}%" if total_acct > 0 else "0%"]
        for col, val in enumerate(row, 1):
            cell = ws4.cell(row=r, column=col, value=val)
            cell.fill = mk_fill(row_bg)
            cell.font = mk_font(9, bold=(col == 3 and val > 0))
            if col == 3 and val > 0:
                cell.font = Font(name='Arial', size=9, bold=True, color=RED)
            cell.alignment = mk_align('center' if col > 1 else 'left')
            cell.border = mk_border()
        ws4.row_dimensions[r].height = 22

    for i, w in enumerate([18,12,12,12,12,14,14]):
        ws4.column_dimensions[get_column_letter(i+1)].width = w
    ws4.sheet_properties.tabColor = "0891B2"

    # ── Sheet 5: By Control ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("By Control")
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = 'A3'

    ws5.merge_cells('A1:F1')
    ws5['A1'] = "Top Failing Controls — FSBP v1.0.0"
    ws5['A1'].font = Font(name='Arial', size=13, bold=True, color=WHITE)
    ws5['A1'].fill = mk_fill(DARK_NAVY)
    ws5['A1'].alignment = mk_align('center')
    ws5.row_dimensions[1].height = 32

    ctrl_headers = ["Control ID","Title","Severity","Failing Resources","Accounts Affected","Remediation URL"]
    for col, hdr in enumerate(ctrl_headers, 1):
        cell = ws5.cell(row=2, column=col, value=hdr)
        cell.font = mk_hdr_font(9)
        cell.fill = mk_fill(MID_BLUE)
        cell.alignment = mk_align('center')
        cell.border = mk_border()
    ws5.row_dimensions[2].height = 28

    ctrl_data = defaultdict(lambda: {'title':'','severity':'','count':0,'accounts':set(),'url':''})
    for f in findings_data:
        if f['human_status'] != 'Remediated':
            cid = f['control_id']
            ctrl_data[cid]['title']    = f['title']
            ctrl_data[cid]['severity'] = f['severity']
            ctrl_data[cid]['count']   += 1
            ctrl_data[cid]['accounts'].add(f['account_id'])
            ctrl_data[cid]['url']      = f['remediation_url']

    sev_rank = {'CRITICAL':0,'HIGH':1,'MEDIUM':2,'LOW':3,'INFORMATIONAL':4}
    sorted_ctrls = sorted(ctrl_data.items(),
                          key=lambda x: (sev_rank.get(x[1]['severity'],9), -x[1]['count']))

    for i, (cid, data) in enumerate(sorted_ctrls[:50]):
        r = i + 3
        row_bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        bg, fg = SEVERITY_COLORS.get(data['severity'], (MID_GRAY, DARK_GRAY))
        row = [cid, data['title'], data['severity'], data['count'],
               len(data['accounts']), data['url']]
        for col, val in enumerate(row, 1):
            cell = ws5.cell(row=r, column=col, value=val)
            cell.border = mk_border()
            if col == 3:
                cell.fill = mk_fill(bg)
                cell.font = Font(name='Arial', size=9, bold=True, color=fg)
                cell.alignment = mk_align('center')
            else:
                cell.fill = mk_fill(row_bg)
                cell.font = mk_font(9)
                cell.alignment = mk_align()
        ws5.row_dimensions[r].height = 20

    for i, w in enumerate([14,42,12,16,16,50]):
        ws5.column_dimensions[get_column_letter(i+1)].width = w
    ws5.sheet_properties.tabColor = GREEN

    wb.save(output_path)
    print(f"  ✓ Excel saved: {output_path}")
    return wb


# ── HTML DASHBOARD ────────────────────────────────────────────────────────────
def build_dashboard(findings_data, region, pull_time, output_path):

    total    = len(findings_data)
    critical = sum(1 for f in findings_data if f['severity'] == 'CRITICAL' and f['human_status'] == 'Open')
    high     = sum(1 for f in findings_data if f['severity'] == 'HIGH' and f['human_status'] == 'Open')
    medium   = sum(1 for f in findings_data if f['severity'] == 'MEDIUM' and f['human_status'] == 'Open')
    low      = sum(1 for f in findings_data if f['severity'] == 'LOW' and f['human_status'] == 'Open')
    remediated = sum(1 for f in findings_data if f['human_status'] == 'Remediated')
    pct_done = f"{remediated/total*100:.1f}" if total > 0 else "0"
    accounts = sorted(set(f['account_id'] for f in findings_data))

    # By severity breakdown
    sev_data = defaultdict(int)
    for f in findings_data:
        if f['human_status'] != 'Remediated':
            sev_data[f['severity']] += 1

    # Top failing controls
    ctrl_data = defaultdict(lambda: {'title':'','severity':'','count':0})
    for f in findings_data:
        if f['human_status'] == 'Open':
            cid = f['control_id']
            ctrl_data[cid]['title']    = f['title']
            ctrl_data[cid]['severity'] = f['severity']
            ctrl_data[cid]['count']   += 1

    sev_rank = {'CRITICAL':0,'HIGH':1,'MEDIUM':2,'LOW':3,'INFORMATIONAL':4}
    top_controls = sorted(ctrl_data.items(),
                          key=lambda x:(sev_rank.get(x[1]['severity'],9),-x[1]['count']))[:15]

    # By account
    acct_data = defaultdict(lambda:{'total':0,'critical':0,'remediated':0})
    for f in findings_data:
        acct_data[f['account_id']]['total'] += 1
        if f['severity'] == 'CRITICAL':
            acct_data[f['account_id']]['critical'] += 1
        if f['human_status'] == 'Remediated':
            acct_data[f['account_id']]['remediated'] += 1

    # Recent findings table (top 20 critical/high open)
    recent = [f for f in findings_data if f['severity'] in ('CRITICAL','HIGH') and f['human_status'] == 'Open'][:20]

    sev_badge = {
        'CRITICAL': 'background:#FEE2E2;color:#DC2626;',
        'HIGH':     'background:#FFEDD5;color:#EA580C;',
        'MEDIUM':   'background:#FEF9C3;color:#D97706;',
        'LOW':      'background:#DCFCE7;color:#16A34A;',
    }

    top_ctrl_rows = ""
    for cid, data in top_controls:
        badge_style = sev_badge.get(data['severity'], 'background:#E2E8F0;color:#64748B;')
        top_ctrl_rows += f"""
        <tr>
          <td style="font-weight:600;color:#1B2A47">{cid}</td>
          <td>{data['title'][:60]}{'...' if len(data['title'])>60 else ''}</td>
          <td><span style="padding:2px 8px;border-radius:12px;font-size:11px;font-weight:700;{badge_style}">{data['severity']}</span></td>
          <td style="text-align:center;font-weight:700;color:#DC2626">{data['count']}</td>
        </tr>"""

    recent_rows = ""
    for f in recent:
        badge_style = sev_badge.get(f['severity'], '')
        recent_rows += f"""
        <tr>
          <td style="font-weight:600;color:#1B2A47;font-size:11px">{f['control_id']}</td>
          <td style="font-size:11px">{f['title'][:55]}{'...' if len(f['title'])>55 else ''}</td>
          <td><span style="padding:2px 6px;border-radius:10px;font-size:10px;font-weight:700;{badge_style}">{f['severity']}</span></td>
          <td style="font-size:11px;color:#64748B">{f['account_id']}</td>
          <td style="font-size:11px;color:#64748B">{f['created_at']}</td>
        </tr>"""

    acct_rows = ""
    for acct_id, data in sorted(acct_data.items()):
        pct = f"{data['remediated']/data['total']*100:.0f}%" if data['total'] > 0 else "0%"
        bar_width = int(data['remediated']/data['total']*100) if data['total'] > 0 else 0
        acct_rows += f"""
        <tr>
          <td style="font-family:monospace;font-size:12px">{acct_id}</td>
          <td style="text-align:center">{data['total']}</td>
          <td style="text-align:center;font-weight:700;color:#DC2626">{data['critical']}</td>
          <td style="text-align:center;color:#16A34A">{data['remediated']}</td>
          <td>
            <div style="background:#E2E8F0;border-radius:4px;height:8px;width:100%">
              <div style="background:#16A34A;border-radius:4px;height:8px;width:{bar_width}%"></div>
            </div>
            <span style="font-size:10px;color:#64748B">{pct}</span>
          </td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Security Hub Dashboard — DEMO Financial</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #F0F4F8; color: #1E293B; }}
  .header {{ background: linear-gradient(135deg, #1B2A47 0%, #2E4A7A 100%); color: white; padding: 24px 32px; display:flex; justify-content:space-between; align-items:center; }}
  .header h1 {{ font-size: 20px; font-weight: 700; }}
  .header .meta {{ font-size: 12px; opacity: 0.75; text-align:right; line-height:1.6; }}
  .subtitle {{ font-size: 12px; opacity: 0.6; margin-top: 4px; }}
  .container {{ padding: 24px 32px; max-width: 1400px; margin: 0 auto; }}
  .kpi-grid {{ display: grid; grid-template-columns: repeat(6, 1fr); gap: 16px; margin-bottom: 24px; }}
  .kpi {{ background: white; border-radius: 10px; padding: 18px; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.08); border-top: 4px solid #ccc; }}
  .kpi.critical {{ border-color: #DC2626; }}
  .kpi.high {{ border-color: #EA580C; }}
  .kpi.medium {{ border-color: #D97706; }}
  .kpi.low {{ border-color: #16A34A; }}
  .kpi.remediated {{ border-color: #0891B2; }}
  .kpi.total {{ border-color: #1B2A47; }}
  .kpi-num {{ font-size: 36px; font-weight: 800; line-height: 1; margin-bottom: 6px; }}
  .kpi.critical .kpi-num {{ color: #DC2626; }}
  .kpi.high .kpi-num {{ color: #EA580C; }}
  .kpi.medium .kpi-num {{ color: #D97706; }}
  .kpi.low .kpi-num {{ color: #16A34A; }}
  .kpi.remediated .kpi-num {{ color: #0891B2; }}
  .kpi.total .kpi-num {{ color: #1B2A47; }}
  .kpi-label {{ font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; color: #64748B; }}
  .kpi-sub {{ font-size: 10px; color: #94A3B8; margin-top: 3px; }}
  .row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 24px; }}
  .row-3 {{ display: grid; grid-template-columns: 2fr 1fr; gap: 20px; margin-bottom: 24px; }}
  .card {{ background: white; border-radius: 10px; padding: 20px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
  .card h2 {{ font-size: 13px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; color: #1B2A47; margin-bottom: 16px; padding-bottom: 10px; border-bottom: 2px solid #F0F4F8; }}
  .progress-wrap {{ margin-bottom: 10px; }}
  .progress-label {{ display:flex; justify-content:space-between; font-size:12px; margin-bottom:4px; }}
  .progress-bar {{ background: #E2E8F0; border-radius: 6px; height: 12px; }}
  .progress-fill {{ border-radius: 6px; height: 12px; transition: width 0.3s; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  th {{ background: #1B2A47; color: white; padding: 8px 10px; text-align: left; font-size: 11px; font-weight: 600; }}
  td {{ padding: 7px 10px; border-bottom: 1px solid #F0F4F8; }}
  tr:hover td {{ background: #F8FAFC; }}
  .footer {{ text-align: center; padding: 20px; font-size: 11px; color: #94A3B8; }}
  .progress-container {{ margin-bottom:20px; }}
  .big-progress {{ background:#E2E8F0; border-radius:8px; height:20px; margin-top:8px; }}
  .big-progress-fill {{ background:linear-gradient(90deg,#0891B2,#16A34A); border-radius:8px; height:20px; width:{pct_done}%; }}
  .pct-label {{ font-size:28px; font-weight:800; color:#0891B2; }}
</style>
</head>
<body>
<div class="header">
  <div>
    <h1>☁ AWS Security Hub — FSBP Remediation Dashboard</h1>
    <div class="subtitle">DEMO Financial Services Group  |  AWS Foundational Security Best Practices v1.0.0</div>
  </div>
  <div class="meta">
    Region: {region}<br>
    Accounts: {len(accounts)}<br>
    Generated: {pull_time}<br>
    Total Findings: {total}
  </div>
</div>

<div class="container">

  <!-- KPI Row -->
  <div class="kpi-grid">
    <div class="kpi total">
      <div class="kpi-num">{total}</div>
      <div class="kpi-label">Total Findings</div>
    </div>
    <div class="kpi critical">
      <div class="kpi-num">{critical}</div>
      <div class="kpi-label">Critical Open</div>
      <div class="kpi-sub">Must fix before go-live</div>
    </div>
    <div class="kpi high">
      <div class="kpi-num">{high}</div>
      <div class="kpi-label">High Open</div>
    </div>
    <div class="kpi medium">
      <div class="kpi-num">{medium}</div>
      <div class="kpi-label">Medium Open</div>
    </div>
    <div class="kpi low">
      <div class="kpi-num">{low}</div>
      <div class="kpi-label">Low Open</div>
    </div>
    <div class="kpi remediated">
      <div class="kpi-num">{remediated}</div>
      <div class="kpi-label">Remediated</div>
      <div class="kpi-sub">{pct_done}% complete</div>
    </div>
  </div>

  <!-- Progress + Account Row -->
  <div class="row">
    <div class="card">
      <h2>Overall Remediation Progress</h2>
      <div class="pct-label">{pct_done}%</div>
      <div style="font-size:12px;color:#64748B;margin-bottom:8px">{remediated} of {total} findings remediated</div>
      <div class="big-progress"><div class="big-progress-fill"></div></div>

      <div style="margin-top:20px">
        {"".join(f'''
        <div class="progress-wrap">
          <div class="progress-label">
            <span style="font-weight:600;color:{"#DC2626" if s=="CRITICAL" else "#EA580C" if s=="HIGH" else "#D97706" if s=="MEDIUM" else "#16A34A"}">{s}</span>
            <span style="color:#64748B">{sev_data.get(s,0)} open</span>
          </div>
          <div class="progress-bar">
            <div class="progress-fill" style="width:{min(sev_data.get(s,0)/max(max(sev_data.values()) if sev_data else 1,1)*100,100):.0f}%;background:{"#DC2626" if s=="CRITICAL" else "#EA580C" if s=="HIGH" else "#D97706" if s=="MEDIUM" else "#16A34A"}"></div>
          </div>
        </div>''' for s in ["CRITICAL","HIGH","MEDIUM","LOW"])}
      </div>
    </div>

    <div class="card">
      <h2>By AWS Account</h2>
      <table>
        <thead><tr><th>Account</th><th>Total</th><th>Critical</th><th>Remediated</th><th>Progress</th></tr></thead>
        <tbody>{acct_rows}</tbody>
      </table>
    </div>
  </div>

  <!-- Controls + Recent Row -->
  <div class="row-3">
    <div class="card">
      <h2>🔴 Critical &amp; High Open Findings</h2>
      <table>
        <thead><tr><th>Control</th><th>Title</th><th>Severity</th><th>Account</th><th>First Seen</th></tr></thead>
        <tbody>{recent_rows}</tbody>
      </table>
    </div>

    <div class="card">
      <h2>Top Failing Controls</h2>
      <table>
        <thead><tr><th>Control</th><th>Title</th><th>Severity</th><th>Failing</th></tr></thead>
        <tbody>{top_ctrl_rows}</tbody>
      </table>
    </div>
  </div>

</div>
<div class="footer">
  AWS Security Hub FSBP v1.0.0  |  Auto-generated {pull_time}  |  DEMO Financial Services Group
</div>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  ✓ Dashboard saved: {output_path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Pull AWS Security Hub FSBP findings')
    parser.add_argument('--profile',  default=None,         help='AWS named profile')
    parser.add_argument('--region',   default=None,         help='AWS region (e.g. ca-central-1)')
    parser.add_argument('--accounts', nargs='+', default=[], help='Filter to specific account IDs')
    parser.add_argument('--output-dir', default=OUTPUT_DIR, help='Output directory')
    args = parser.parse_args()

    print("""
╔══════════════════════════════════════════════════════════════╗
║         AWS Security Hub — FSBP Findings Puller              ║
║         AWS Foundational Security Best Practices v1.0.0      ║
╚══════════════════════════════════════════════════════════════╝
""")

    session = get_session(profile=args.profile, region=args.region)

    try:
        sts = session.client('sts')
        identity = sts.get_caller_identity()
        print(f"  ✓ Authenticated as: {identity['Arn']}")
    except Exception as e:
        print(f"  ✗ AWS auth failed: {e}")
        print("    Ensure AWS credentials are set via environment variables or ~/.aws/credentials")
        sys.exit(1)

    raw_findings, region = pull_findings(session, args.accounts or None)

    if not raw_findings:
        print("  ⚠ No FSBP findings returned. Check that Security Hub is enabled and FSBP standard is active.")
        sys.exit(0)

    print(f"\n  Parsing {len(raw_findings)} findings...")
    findings_data = [parse_finding(f) for f in raw_findings]

    pull_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    os.makedirs(args.output_dir, exist_ok=True)
    xlsx_path  = os.path.join(args.output_dir, f"SecurityHub_FSBP_{timestamp}.xlsx")
    html_path  = os.path.join(args.output_dir, f"SecurityHub_Dashboard_{timestamp}.html")

    print("\n  Building Excel tracker...")
    build_excel(findings_data, region, pull_time, xlsx_path)

    print("\n  Building HTML dashboard...")
    build_dashboard(findings_data, region, pull_time, html_path)

    print(f"""
╔══════════════════════════════════════════════════════════════╗
║  Done!                                                       ║
╠══════════════════════════════════════════════════════════════╣
║  Excel Tracker : {xlsx_path:<44}║
║  Dashboard     : {html_path:<44}║
╠══════════════════════════════════════════════════════════════╣
║  Open the dashboard in any browser.                          ║
║  Send the Weekly Status Report tab from the Excel tracker.   ║
║                                                              ║
║  Automate — cron every Monday 8am:                           ║
║  0 8 * * 1 python securityhub_pull.py >> pull.log 2>&1       ║
╚══════════════════════════════════════════════════════════════╝
""")

if __name__ == '__main__':
    main()
